#################################################
# 해결 1. 텍스트 목록(제목을 누르면 텍스트를 읽을 수 있는)은 구글이나 네이버와 거의 비슷한 비쥬얼이면 좋을 것 같고
# 해결 2. 텍스트 목록은 제목, 짧은 텍스트 내용 일부
# 3. 텍스트 목록의 제목은 클릭을 하고나면 읽었다는게 표시되도록 기존 포털처럼 누른건 색상 변화가 있었으면 좋을 것 같고
# 해결 4. 텍스트 목록의 텍스트들은 사람들에게 랜덤르로 제시되어야 하고(이게 포인트: 텍스트 순서가 읽기에 영향을 미치기 때문에)
# 5. 텍스트 이동 기록은 시간순으로 나열되고, 머문 시간도 기록되어야 하고
#  6. 메모프로그램은 어떤 글을 보고 어떤 내용을 썼는지 분석하기 위함
# 7. 이 부분이 새로 생긴 부분인데, 텍스트를 하나 읽고 다시 텍스트 목록으로 돌아갈때, 텍스트로 다시 돌아가기, 읽기를 끝내기 버튼이 있는데 이때 텍스트로 다시 돌아가기를 누르면 질문이 나오고 그에 대한 대답을 해야 다시 텍스트 목록이 나오도록 설정되어야해.
# 8. 텍스트 내부에 글을 쓴 사람 이름(이름 아니더라도) 누르면 작게 창이 뜨면서(꼭 창이 아니어도 되고) 정보가 제공되어야 하고 이 버튼을 눌렀는지 누르지 않았는지에 대한 기록이 필요
#################################################

import tkinter as tk
import time
from openpyxl import Workbook, load_workbook
import random

class ButtonApp:
    def __init__(self, root):
        self.root = root
        self.root.title("버튼 클릭 타이머")
        self.root.geometry("800x1000")

        self.button_info = {
            0: {
                "text": "속보: 대규모 이벤트 발생 중\n주요 내용 갱신 중",
                "author": {
                    "name": "작가1",
                    "occupation": "작가",
                    "email": "author1@email.com"
                }
            },
            1: {
                "text": "주요 이슈: 단독 인터뷰\n최신 업데이트",
                "author": {
                    "name": "작가2",
                    "occupation": "기자",
                    "email": "author2@email.com"
                }
            },
            2: {
                "text": "최신 업데이트: 시장 동향\n기술 혁신 소식",
                "author": {
                    "name": "작가3",
                    "occupation": "편집자",
                    "email": "author3@email.com"
                }
            },
            3: {
                "text": "기술 혁신 소식: 혁신성 제품\n기술 혁신",
                "author": {
                    "name": "작가4",
                    "occupation": "논설위원",
                    "email": "author4@email.com"
                }
            },
            4: {
                "text": "건강과 웰빙: 전문가 인사이트\n건강 관련 소식",
                "author": {
                    "name": "작가5",
                    "occupation": "기술분석가",
                    "email": "author5@email.com"
                }
            },
            5: {
                "text": "예술과 문화: 아티스트 초점\n문화 소식",
                "author": {
                    "name": "작가6",
                    "occupation": "예술가",
                    "email": "author6@email.com"
                }
            },
            6: {
                "text": "과학과 환경: 최근 연구 발표\n과학 소식",
                "author": {
                    "name": "작가7",
                    "occupation": "과학자",
                    "email": "author7@email.com"
                }
            },
            7: {
                "text": "스포츠 하이라이트: 짜릿한 경기\n스포츠 소식",
                "author": {
                    "name": "작가8",
                    "occupation": "스포츠 해설가",
                    "email": "author8@email.com"
                }
            }
        }

        self.buttons = []
        self.button_start_times = {}
        self.log_file = "log.xlsx"

        self.load_log_file()
        self.root.bind('<Escape>', self.exit_program)

    def create_buttons(self):
        button_indices = list(range(8))
        random.shuffle(button_indices)

        # Create a shuffled list of button text
        button_texts = [self.button_info[i]["text"] for i in button_indices]

        for i, text in enumerate(button_texts):
            button_info = self.button_info[button_indices[i]]
            button = tk.Button(self.root, text=text, width=50, height=5, font=("Arial", 12),
                               command=lambda i=button_indices[i]: self.on_button_click(i))
            button.grid(row=i, column=0, padx=10, pady=10, sticky='nsew')
            self.buttons.append(button)

        for i in range(8):
            self.root.grid_rowconfigure(i, weight=1)
            self.root.grid_columnconfigure(0, weight=1)

    def load_log_file(self):
        try:
            self.log_workbook = load_workbook(self.log_file)
        except FileNotFoundError:
            self.log_workbook = Workbook()
            self.log_workbook.save(self.log_file)

        if "Log" in self.log_workbook.sheetnames:
            self.log_sheet = self.log_workbook["Log"]
        else:
            self.log_sheet = self.log_workbook.active
            self.log_sheet.title = "Log"
            self.log_sheet.append(["Button", "Timestamp", "Duration"])

    def on_button_click(self, button_index):
        self.button_start_times[button_index] = time.time()
        self.create_text_window(button_index)

    def create_text_window(self, button_index):
        text_window = tk.Toplevel(self.root)
        text_window.title(f"뉴스 헤드라인 - 버튼 {button_index + 1}")

        text_widget = tk.Text(text_window, wrap=tk.WORD, width=60, height=20)
        text_widget.pack()

        button_info = self.button_info[button_index]
        text_widget.insert(tk.END, button_info["text"])

        author_info_button = tk.Button(text_window, text="저자 정보",
                                       command=lambda i=button_index: self.show_author_info(i))
        author_info_button.pack()

        close_button = tk.Button(text_window, text="닫기", command=text_window.destroy)
        close_button.pack()

    def show_author_info(self, button_index):
        author_info_window = tk.Toplevel(self.root)
        author_info_window.title("저자 정보")

        button_info = self.button_info[button_index]
        author_info = button_info["author"]

        author_info_text = f"저자: {author_info['name']}\n직업: {author_info['occupation']}\n이메일: {author_info['email']}"

        author_info_label = tk.Label(author_info_window, text=author_info_text)
        author_info_label.pack()

    def exit_program(self, event):
        self.root.quit()

if __name__ == "__main__":
    root = tk.Tk()
    app = ButtonApp(root)
    app.create_buttons()
    root.mainloop()
